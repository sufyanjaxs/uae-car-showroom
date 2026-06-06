import io
import csv
from typing import List, Dict, Any
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


def export_to_csv(data: List[Dict[str, Any]], filename: str = "export.csv") -> StreamingResponse:
    output = io.StringIO()
    if data:
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def export_to_excel(data: List[Dict[str, Any]], filename: str = "export.xlsx") -> StreamingResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    if data:
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for col_idx, key in enumerate(data[0].keys(), 1):
            cell = ws.cell(row=1, column=col_idx, value=key)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        for row_idx, row in enumerate(data, 2):
            for col_idx, key in enumerate(row.keys(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row[key])
                cell.border = thin_border
        ws.column_dimensions["A"].width = 30
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def export_to_pdf(data: List[Dict[str, Any]], filename: str = "export.pdf") -> StreamingResponse:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    if data:
        table_data = [list(data[0].keys())]
        for row in data:
            table_data.append([str(v) for v in row.values()])
        table = Table(table_data)
        style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ])
        table.setStyle(style)
        elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
